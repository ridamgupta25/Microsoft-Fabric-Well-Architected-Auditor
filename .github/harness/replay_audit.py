#!/usr/bin/env python
r"""Replay an audit report against its archived KB snapshot — the mechanical oracle.

Run from ``backend/`` with the venv interpreter::

    ..\.venv\Scripts\python.exe ..\.github\harness\replay_audit.py

Scoring is a pure function of the crawled :class:`WorkspaceContext`, so the
report in ``output/`` must be exactly reproducible from the ``workspace.json``
archived for the same run. This re-runs the real engine over the archived
snapshot and diffs every recomputed verdict against the reported row —
check id, object, status, score, and the evidence string.

That answers a question no amount of reading can: *does this report actually
correspond to the data that was captured?* A clean replay proves the report is
faithful to the snapshot and that the checks are deterministic. It does **not**
prove the snapshot is a complete picture of the tenant (see the crawl-
completeness counters in ``summary.json``) nor that a check's rule is the right
rule — those stay human judgement.

Exit code is 0 only when every row matches.
"""
from __future__ import annotations

import argparse
import json
import sys
from collections import Counter
from collections.abc import Iterable
from pathlib import Path

import auditfast.core.check  # noqa: F401 - importing the package registers every check
from auditfast.config.settings import get_settings
from auditfast.core.engine import READ_INCOMPLETE_CHECK_ID
from auditfast.core.engine import run_audit as run_engine
from auditfast.core.enums import Layer, Resource
from auditfast.core.models import WorkspaceContext
from auditfast.services.audit_service import ACCESS_CHECK_ID
from auditfast.services.project import load_project, load_remediation

#: Rows the audit routes to ``errors[]`` rather than the Checks sheet.
_NON_CHECK_IDS = {ACCESS_CHECK_ID, READ_INCOMPLETE_CHECK_ID}

#: One comparable row: what the report claims / what the replay computes.
Row = tuple[str, str, str, str, str, str]  # workspace, check_id, obj, status, score, evidence


def _norm(value: object) -> str:
    """Normalize a cell for comparison.

    Both sides go through this. Fabric item names routinely carry stray leading
    or trailing spaces, and Excel round-trips them faithfully — normalizing only
    one side would report drift where there is none.
    """
    return "" if value is None else str(value).strip()


class SnapshotProvider:
    """Serves archived ``workspace.json`` snapshots. No network, no cache."""

    def __init__(self, snapshots: dict[str, dict]):
        self._snapshots = snapshots

    def fetch(
        self,
        workspace_id: str,
        layer: Layer = Layer.MIXED,
        resources: Iterable[Resource] | None = None,
    ) -> WorkspaceContext:
        # A fresh context per call: a check must never see another run's mutations.
        return WorkspaceContext.from_dict(self._snapshots[workspace_id])

    def list_workspaces(self) -> list[dict]:
        return []


def _latest_snapshots(archive_root: Path) -> dict[str, tuple[Path, dict]]:
    """Newest ``workspace.json`` per workspace folder, keyed by display name."""
    found: dict[str, tuple[Path, dict]] = {}
    for workspace_dir in sorted(p for p in archive_root.iterdir() if p.is_dir()):
        candidates = sorted(
            workspace_dir.glob("*/workspace.json"),
            key=lambda p: p.stat().st_mtime,
        )
        if not candidates:
            continue
        newest = candidates[-1]
        data = json.loads(newest.read_text(encoding="utf-8"))
        found[data.get("display_name") or data["id"]] = (newest, data)
    return found


def _report_rows(report: Path) -> list[Row]:
    """Every data row of the report's Checks sheet, normalized for comparison."""
    from openpyxl import load_workbook

    workbook = load_workbook(report, read_only=True, data_only=True)
    if "Checks" not in workbook.sheetnames:
        raise SystemExit(f"{report} has no 'Checks' sheet — is it an audit report?")
    sheet = workbook["Checks"]
    rows = sheet.iter_rows(values_only=True)
    header = [str(h or "") for h in next(rows)]
    index = {name: i for i, name in enumerate(header)}
    required = ("Workspace", "Object", "Check ID", "Status", "Score", "Evidence")
    missing = [name for name in required if name not in index]
    if missing:
        raise SystemExit(f"Checks sheet is missing column(s): {', '.join(missing)}")

    out: list[Row] = []
    for raw in rows:
        if raw is None or all(cell is None for cell in raw):
            continue
        out.append(tuple(  # type: ignore[arg-type]
            _norm(raw[index[name]])
            for name in ("Workspace", "Check ID", "Object", "Status", "Score", "Evidence")
        ))
    workbook.close()
    return out


def _replay_rows(snapshots: dict[str, tuple[Path, dict]], project_path: Path) -> list[Row]:
    """Re-run the engine over the snapshots and return comparable rows."""
    config = load_project(project_path)
    by_id = {data["id"]: data for _, data in snapshots.values()}
    targets = [
        (ws_id, Layer(data["layer"]) if data.get("layer") else Layer.MIXED)
        for ws_id, data in by_id.items()
    ]
    results = run_engine(
        SnapshotProvider(by_id), targets, config.settings,
        remediation=load_remediation(config),
    )
    return [
        (_norm(r.workspace), _norm(r.check_id), _norm(r.obj), _norm(r.status.value),
         "" if r.score is None else str(r.score), _norm(r.evidence))
        for r in results
        if r.check_id not in _NON_CHECK_IDS
    ]


def _describe(row: Row) -> str:
    workspace, check_id, obj, status, score, _ = row
    where = f"{workspace}/{obj}" if obj else workspace
    return f"{check_id} @ {where} -> {status} ({score or 'unscored'})"


def compare(reported: list[Row], replayed: list[Row]) -> dict:
    """Diff the two row sets, classifying every difference by its likely cause."""
    # Key on identity; the verdict fields are what we are testing.
    def keyed(rows: list[Row]) -> dict[tuple[str, str, str], list[Row]]:
        out: dict[tuple[str, str, str], list[Row]] = {}
        for row in rows:
            out.setdefault(row[:3], []).append(row)
        return out

    left, right = keyed(reported), keyed(replayed)
    matched: list[Row] = []
    mismatched: list[dict] = []

    for key in left.keys() & right.keys():
        # Same identity may legitimately appear more than once; compare as multisets.
        want, got = Counter(left[key]), Counter(right[key])
        matched.extend((want & got).elements())
        for row in (want - got).elements():
            partner = next(iter((got - want).elements()), None)
            reasons = []
            if partner is None:
                reasons.append("no recomputed verdict left to pair with")
            else:
                if row[3] != partner[3]:
                    reasons.append(f"status: report {row[3]} vs replay {partner[3]}")
                if row[4] != partner[4]:
                    want_score = row[4] or "unscored"
                    got_score = partner[4] or "unscored"
                    reasons.append(f"score: report {want_score} vs replay {got_score}")
                if row[5] != partner[5]:
                    reasons.append(f"evidence:\n      report: {row[5]}\n      replay: {partner[5]}")
            mismatched.append({"row": _describe(row), "reasons": reasons})

    only_report = [r for key in left.keys() - right.keys() for r in left[key]]
    only_replay = [r for key in right.keys() - left.keys() for r in right[key]]
    return {
        "matched": matched,
        "mismatched": mismatched,
        "only_in_report": only_report,
        "only_in_replay": only_replay,
    }


def main(argv: list[str]) -> int:
    settings = get_settings()
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--report", default=str(settings.output_path / "audit-report.xlsx"),
                        help="the audit report to verify")
    parser.add_argument("--archive", default=str(settings.resolve(settings.kb_archive_dir)),
                        help="root of the timestamped KB archive")
    parser.add_argument("--project", default=str(settings.project_path),
                        help="the project YAML the audit ran with (its thresholds "
                             "and naming regexes must match, or checks will differ)")
    parser.add_argument("--json", action="store_true", help="emit machine-readable output")
    args = parser.parse_args(argv)

    report = Path(args.report)
    archive = Path(args.archive)
    if not report.exists():
        print(f"FAIL  no report at {report} — run an audit first.")
        return 2
    if not archive.exists():
        print(f"FAIL  no KB archive at {archive} — is AUDITFAST_KB_ARCHIVE_ENABLED off?")
        return 2

    snapshots = _latest_snapshots(archive)
    if not snapshots:
        print(f"FAIL  {archive} holds no workspace.json snapshots.")
        return 2

    reported = _report_rows(report)
    # Replay only what the report covers: an unrelated snapshot on disk is not
    # evidence about this report, and replaying it would drown the real diff.
    covered = {row[0] for row in reported}
    orphaned = sorted(covered - snapshots.keys())
    if orphaned:
        print("FAIL  the report covers workspace(s) with no archived snapshot: "
              f"{', '.join(orphaned)}")
        print(f"      looked under {archive}\\<workspace name>\\*\\workspace.json")
        return 2
    snapshots = {name: entry for name, entry in snapshots.items() if name in covered}

    replayed = _replay_rows(snapshots, Path(args.project))
    result = compare(reported, replayed)

    if args.json:
        print(json.dumps({
            "report": str(report),
            "snapshots": {name: str(path) for name, (path, _) in snapshots.items()},
            "matched": len(result["matched"]),
            "mismatched": result["mismatched"],
            "only_in_report": [_describe(r) for r in result["only_in_report"]],
            "only_in_replay": [_describe(r) for r in result["only_in_replay"]],
        }, indent=2))
        return 0 if _clean(result) else 1

    print(f"report    {report}")
    print(f"project   {args.project}")
    for name, (path, _) in sorted(snapshots.items()):
        print(f"snapshot  {name}  {path}")
    print()
    print(f"OK    {len(result['matched'])} of {len(reported)} reported rows reproduced "
          f"exactly from the snapshot")

    for entry in result["mismatched"]:
        print(f"FAIL  {entry['row']}")
        for reason in entry["reasons"]:
            print(f"      {reason}")
    for row in result["only_in_report"]:
        print(f"FAIL  in report but not reproducible: {_describe(row)}")
        print("      the report predates this snapshot, or the check was removed/renamed")
    for row in result["only_in_replay"]:
        print(f"WARN  produced by replay but absent from the report: {_describe(row)}")
        print("      the audit likely ran with a pillar filter, or the check is new "
              "since the report was written")

    if _clean(result):
        print("\nGO    the report is faithful to the captured snapshot and the checks "
              "are deterministic.")
        print("      Still unproven by this replay: whether the snapshot itself is a "
              "complete view of the tenant (check summary.json read_failures), and "
              "whether each check's rule is the right rule.")
        return 0
    print("\nNO-GO the report does not correspond to the snapshot — do not trust its "
          "numbers until the differences above are explained.")
    return 1


def _clean(result: dict) -> bool:
    return not result["mismatched"] and not result["only_in_report"]


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
