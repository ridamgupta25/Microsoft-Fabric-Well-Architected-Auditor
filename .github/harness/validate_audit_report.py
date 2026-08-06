#!/usr/bin/env python3
"""Validate a generated audit report against the snapshot knowledge base.

This script reads the generated report from backend/output/audit-report.xlsx (or
backend/output/audit-report.csv), resolves the latest workspace snapshot under
backend/Fabric workspace kb/, and checks whether the KB has enough artifact
metadata to support each finding. It writes a final validation report to
backend/validation/validation-report.md.
"""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("openpyxl is required to run this script. Install it in the backend venv.") from exc

REPO_ROOT = Path(__file__).resolve().parents[2]
BACKEND_ROOT = REPO_ROOT / "backend"
OUTPUT_DIR = BACKEND_ROOT / "output"
KB_ROOT = BACKEND_ROOT / "Fabric workspace kb"
VALIDATION_DIR = BACKEND_ROOT / "validation"
VALIDATION_PATH = VALIDATION_DIR / "validation-report.md"

CHECKS_SHEET_NAME = "Checks"


def normalize_name(value: str | None) -> str:
    if not value:
        return ""
    return value.strip().lower().replace("_", " ").replace("\n", " ")


def load_audit_report() -> list[dict[str, Any]]:
    xlsx_path = OUTPUT_DIR / "audit-report.xlsx"
    csv_path = OUTPUT_DIR / "audit-report.csv"
    if xlsx_path.exists():
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        if CHECKS_SHEET_NAME not in wb.sheetnames:
            raise SystemExit(f"Workbook {xlsx_path} does not contain sheet {CHECKS_SHEET_NAME}")
        sheet = wb[CHECKS_SHEET_NAME]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        header = [cell if cell is not None else "" for cell in rows[0]]
        return [dict(zip(header, row)) for row in rows[1:] if any(cell is not None for cell in row)]
    if csv_path.exists():
        with csv_path.open("r", encoding="utf-8", newline="") as stream:
            reader = csv.DictReader(stream)
            return [row for row in reader]
    raise SystemExit("No audit-report.xlsx or audit-report.csv found in backend/output")


def find_workspace_snapshot(workspace_name: str) -> Path | None:
    normalized_target = normalize_name(workspace_name)
    if not KB_ROOT.exists():
        return None
    candidates = []
    for ws_dir in sorted(KB_ROOT.iterdir()):
        if not ws_dir.is_dir():
            continue
        normalized_dir = normalize_name(ws_dir.name)
        if normalized_dir == normalized_target or normalized_target in normalized_dir or normalized_dir in normalized_target:
            snapshot_dirs = [d for d in ws_dir.iterdir() if d.is_dir()]
            if not snapshot_dirs:
                continue
            latest = max(snapshot_dirs, key=lambda d: d.stat().st_mtime)
            candidates.append((ws_dir.name, latest))
    if not candidates:
        return None
    return candidates[0][1]


def load_snapshot(snapshot_folder: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    summary_path = snapshot_folder / "summary.json"
    workspace_path = snapshot_folder / "workspace.json"
    if not summary_path.exists() or not workspace_path.exists():
        raise SystemExit(f"Missing snapshot files in {snapshot_folder}")
    summary = json.loads(summary_path.read_text(encoding="utf-8"))
    workspace = json.loads(workspace_path.read_text(encoding="utf-8"))
    return summary, workspace


def locate_artifact(workspace: dict[str, Any], object_name: str) -> tuple[str | None, Any | None, str]:
    normalized = normalize_name(object_name)
    for section in ["pipelines", "notebooks", "reports", "semantic_models", "lakehouses", "warehouse", "dataflows", "items"]:
        entry = workspace.get(section)
        if isinstance(entry, dict):
            for key, value in entry.items():
                if normalize_name(key) == normalized:
                    return section, value, key
        elif isinstance(entry, list):
            for item in entry:
                if normalize_name(item.get("display_name")) == normalized:
                    return section, item, item.get("display_name")
    return None, None, object_name


def has_definition(section: str | None, artifact: Any) -> bool:
    if artifact is None:
        return False
    if section == "pipelines":
        return bool(artifact.get("properties") and artifact["properties"].get("activities"))
    if section == "notebooks":
        return bool(artifact.get("cells"))
    if section == "reports":
        return bool(artifact)
    if section in {"semantic_models", "lakehouses", "warehouse", "dataflows", "items"}:
        return True
    return bool(artifact)


def load_workspace_snapshots(rows):
    snapshots = {}
    unknown_workspaces = set()
    for row in rows:
        workspace_name = row.get("Workspace") or row.get("workspace")
        if not workspace_name:
            continue
        normalized = normalize_name(workspace_name)
        if normalized in snapshots or normalized in unknown_workspaces:
            continue
        snapshot_folder = find_workspace_snapshot(workspace_name)
        if snapshot_folder is None:
            unknown_workspaces.add(normalized)
            continue
        summary, workspace = load_snapshot(snapshot_folder)
        snapshots[normalized] = (summary, workspace, snapshot_folder)
    if unknown_workspaces:
        raise SystemExit(
            f"Could not locate snapshots for workspaces: {', '.join(sorted(unknown_workspaces))}"
        )
    return snapshots


def interpret_row(row: dict[str, Any], snapshots: dict[str, tuple[dict[str, Any], dict[str, Any], Path]]) -> dict[str, Any]:
    object_name = row.get("Object") or row.get("object") or ""
    workspace_name = row.get("Workspace") or row.get("workspace")
    if not workspace_name:
        status = "NOT VERIFIABLE"
        reason = "missing_workspace"
        note = "Audit row does not include a workspace name"
        artifact_section = None
        matched_name = object_name
    else:
        normalized_workspace = normalize_name(workspace_name)
        summary, workspace, _ = snapshots[normalized_workspace]
        if not summary.get("complete", False) or summary.get("read_failures"):
            status = "NOT VERIFIABLE"
            reason = "incomplete_capture"
            note = "Workspace knowledge base is incomplete or has read failures"
            artifact_section = None
            matched_name = object_name
        elif not object_name:
            status = "PASS"
            reason = "satisfied"
            note = "Workspace-level check; KB summary is complete and no read failures were recorded"
            artifact_section = "workspace"
            matched_name = "workspace"
        else:
            artifact_section, artifact, matched_name = locate_artifact(workspace, object_name)
            if artifact_section is None:
                status = "NOT VERIFIABLE"
                reason = "missing_artifact"
                note = f"Referenced object '{object_name}' was not found in the KB snapshot"
            elif not has_definition(artifact_section, artifact):
                status = "NOT VERIFIABLE"
                reason = "missing_metadata"
                note = f"Artifact '{matched_name}' is present in section {artifact_section} but lacks definition metadata"
            else:
                status = "PASS"
                reason = "satisfied"
                note = f"KB contains artifact '{matched_name}' in section {artifact_section} with definition metadata"
    return {
        "workspace": workspace_name,
        "check_id": row.get("Check ID") or row.get("Check ID"),
        "ref": row.get("Ref") or row.get("ref"),
        "title": row.get("Title") or row.get("title"),
        "object": object_name,
        "reported_status": row.get("Status"),
        "kb_status": status,
        "reason": reason,
        "note": note,
        "artifact_section": artifact_section or "unknown",
        "matched_name": matched_name,
    }


def build_report(rows: list[dict[str, Any]], snapshots: dict[str, tuple[dict[str, Any], dict[str, Any], Path]]) -> str:
    total = len(rows)
    pass_count = sum(1 for row in rows if row["kb_status"] == "PASS")
    not_verifiable_count = sum(1 for row in rows if row["kb_status"] == "NOT VERIFIABLE")
    workspace_paths = sorted({snapshot[2] for snapshot in snapshots.values()})
    lines = [
        "# Audit KB Validation Report",
        "",
        f"Generated from: `{OUTPUT_DIR / 'audit-report.xlsx'}`",
        "",
        "## Workspace snapshots",
    ]
    for path in workspace_paths:
        lines.append(f"- `{path}`")
    lines.extend([
        "",
        "## Validation summary",
        f"- rows analyzed: `{total}`",
        f"- KB-verifiable: `{pass_count}`",
        f"- NOT VERIFIABLE: `{not_verifiable_count}`",
        "",
        "## Snapshot completeness",
    ])
    for workspace_name, (summary, _, path) in sorted(snapshots.items()):
        lines.append(f"### `{workspace_name}` snapshot: `{path}`")
        lines.append(f"- complete: `{summary.get('complete', False)}`")
        lines.append(f"- notebooks_read: `{summary.get('notebooks_read')}`")
        lines.append(f"- pipelines_read: `{summary.get('pipelines_read')}`")
        lines.append(f"- tables_read: `{summary.get('tables_read')}`")
        lines.append(f"- semantic_models_read: `{summary.get('semantic_models_read')}`")
        lines.append(f"- unavailable: `{summary.get('unavailable', [])}`")
        lines.append(f"- read_failures: `{summary.get('read_failures', {})}`")
        lines.append("")
    lines.extend([
        "## Findings",
        "",
        "| Workspace | Ref | Check ID | Object | Reported | KB coverage | Reason | Note |",
        "|---|---|---|---|---|---|---|---|",
    ])
    for row in rows:
        lines.append(
            f"| {row['workspace'] or ''} | {row['ref'] or ''} | {row['check_id'] or ''} | {row['object'] or ''} | {row['reported_status'] or ''} | {row['kb_status']} | {row['reason']} | {row['note']} |"
        )
    lines.append("")
    lines.append("## KB gaps")
    lines.append("")
    if not_verifiable_count == 0:
        lines.append("No KB gaps detected for the analyzed rows.")
    else:
        for row in rows:
            if row["kb_status"] != "PASS":
                lines.append(f"- `{row['workspace']}` / `{row['ref']}` / `{row['check_id']}` / `{row['object']}`: {row['note']}")
    lines.append("")
    lines.append("## Notes")
    lines.append("- This report assesses whether the knowledge base contains the artifacts and definitions needed to validate each finding, not whether the audit verdict itself is logically correct.")
    return "\n".join(lines)


def main() -> int:
    rows = load_audit_report()
    if not rows:
        raise SystemExit("No rows found in the audit report")
    snapshots = load_workspace_snapshots(rows)
    interpreted = [interpret_row(row, snapshots) for row in rows]
    VALIDATION_DIR.mkdir(parents=True, exist_ok=True)
    VALIDATION_PATH.write_text(build_report(interpreted, snapshots), encoding="utf-8")
    print(f"Validation report written to {VALIDATION_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
