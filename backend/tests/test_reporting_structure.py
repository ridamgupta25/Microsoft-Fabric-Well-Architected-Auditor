from openpyxl import load_workbook

from auditfast.core.enums import Layer, Pillar, Scope, Severity, Status
from auditfast.core.models import CheckResult
from auditfast.core.scoring import aggregate
from auditfast.reporting.excel import build_excel
from auditfast.reporting.markdown import build_markdown
from auditfast.reporting.structure import consolidate, findings


def _result(
    *,
    check_id: str,
    ref: str,
    title: str,
    pillar: Pillar,
    status: Status,
    score: int | None,
    obj: str,
    evidence: str,
    recommendation: str = "",
    severity: Severity = Severity.MEDIUM,
    workspace: str = "Workspace A",
) -> CheckResult:
    return CheckResult(
        check_id=check_id,
        ref=ref,
        title=title,
        pillar=pillar,
        status=status,
        score=score,
        evidence=evidence,
        recommendation=recommendation,
        severity=severity,
        workspace=workspace,
        layer=Layer.MIXED,
        obj=obj,
        scope=Scope.NOTEBOOK,
    )


def _sample_results() -> list[CheckResult]:
    shared = {
        "check_id": "NB-ACCESS",
        "ref": "6.1.1",
        "title": "Notebook access follows least privilege",
        "pillar": Pillar.SECURITY,
        "severity": Severity.HIGH,
        "recommendation": "Restrict notebook access to approved identities.",
    }
    return [
        _result(
            **shared,
            status=Status.FAIL,
            score=0,
            obj="Notebook A",
            evidence="Broad access detected",
        ),
        _result(
            **shared,
            status=Status.FAIL,
            score=0,
            obj="Notebook B",
            evidence="Broad access detected",
        ),
        _result(
            **shared,
            status=Status.PASS,
            score=3,
            obj="Notebook C",
            evidence="Access is restricted",
        ),
        _result(
            **shared,
            status=Status.NA,
            score=None,
            obj="Notebook D",
            evidence="Definition could not be read",
        ),
        _result(
            check_id="NB-CACHE",
            ref="3.2.1",
            title="Notebook uses cache selectively",
            pillar=Pillar.PERFORMANCE,
            status=Status.PASS,
            score=3,
            obj="Notebook A",
            evidence="No unnecessary cache calls",
        ),
    ]


def test_consolidates_repeated_asset_failures_by_control():
    controls = consolidate(_sample_results())

    assert len(controls) == 2
    assert len(findings(controls)) == 1
    access = next(control for control in controls if control.check_id == "NB-ACCESS")
    assert access.status is Status.FAIL
    assert "Notebook A" in access.impacted_assets
    assert "Notebook B" in access.impacted_assets
    assert "Notebook C" in access.non_impacted_assets
    assert "Definition could not be read" in access.not_assessed
    assert "Notebook D" in access.not_assessed


def test_excel_matches_sql_report_flow_and_risk_register(tmp_path):
    results = _sample_results()
    output = tmp_path / "fabric-report.xlsx"

    build_excel(str(output), "Fabric Project", aggregate(results), results)

    workbook = load_workbook(output, data_only=True)
    assert workbook.sheetnames == [
        "Summary",
        "Area Detail",
        "Checklist",
        "Findings",
        "Risk Register",
        "Invent",
    ]
    assert workbook["Checklist"].max_row == 3
    assert workbook["Findings"].max_row == 2
    assert workbook["Risk Register"].max_row == 12
    assert workbook["Risk Register"]["A11"].value == "Risk ID"
    assert workbook["Risk Register"]["Z11"].value == "Notes"
    assert "Notebook A" in workbook["Findings"]["F2"].value
    assert "Notebook B" in workbook["Findings"]["F2"].value
    assert "Notebook C" in workbook["Findings"]["G2"].value
    assert workbook["Summary"].freeze_panes == "A3"
    assert workbook["Area Detail"].freeze_panes == "B2"
    assert workbook["Checklist"].freeze_panes == "J2"
    assert workbook["Findings"].freeze_panes == "F2"
    assert workbook["Invent"].freeze_panes == "C2"
    assert set(workbook["Checklist"].tables) == {"ChecklistTable"}
    assert set(workbook["Findings"].tables) == {"FindingsTable"}
    assert set(workbook["Risk Register"].tables) == {"RiskRegisterTable"}
    assert set(workbook["Invent"].tables) == {"InventoryTable"}
    assert workbook["Checklist"].tables["ChecklistTable"].tableStyleInfo.showRowStripes
    assert len(workbook["Checklist"].conditional_formatting) > 0
    assert len(workbook["Summary"].conditional_formatting) > 0
    assert workbook["Checklist"]["A1"].font.bold
    assert workbook["Checklist"]["A1"].fill.fgColor.rgb.endswith("305496")
    assert workbook["Checklist"]["J2"].number_format == "0.00"
    assert 10 <= workbook["Checklist"].column_dimensions["A"].width <= 60


def test_markdown_matches_sql_section_hierarchy():
    results = _sample_results()

    report = build_markdown("Fabric Project", aggregate(results), results)

    headings = [
        "## Executive Summary",
        "## Area Detail",
        "## Checklist",
        "## Findings (1)",
        "## Risk Register",
        "## Invent",
    ]
    positions = [report.index(f"\n{heading}\n") for heading in headings]
    assert positions == sorted(positions)
    assert "### Remediation Roadmap" in report
    assert "Notebook A" in report
    assert "Notebook B" in report
    assert "Notebook C" in report
    assert "R-001" in report


def test_workspace_ids_and_scores_match_between_inventory_and_checklist(tmp_path):
    results = _sample_results()
    results.append(
        _result(
            check_id="NB-CACHE",
            ref="3.2.1",
            title="Notebook uses cache selectively",
            pillar=Pillar.PERFORMANCE,
            status=Status.FAIL,
            score=0,
            obj="Notebook Z",
            evidence="Unnecessary cache calls",
            workspace="Workspace B",
        )
    )
    output = tmp_path / "multi-workspace-report.xlsx"

    build_excel(str(output), "Fabric Project", aggregate(results), results)

    workbook = load_workbook(output, read_only=True, data_only=True)
    inventory = workbook["Invent"]
    checklist = workbook["Checklist"]
    inventory_ids = {
        inventory.cell(row=row, column=2).value: inventory.cell(row=row, column=1).value
        for row in range(2, inventory.max_row + 1)
    }
    assert inventory_ids == {"Workspace A": "WS1", "Workspace B": "WS2"}

    checklist_headers = {
        checklist.cell(row=1, column=column).value: column
        for column in range(1, checklist.max_column + 1)
    }
    assert "Workspace ID" not in checklist_headers
    assert checklist_headers["WS1"] == checklist.max_column - 1
    assert checklist_headers["WS2"] == checklist.max_column
    checklist_rows = {
        checklist.cell(row=row, column=checklist_headers["Check ID"]).value: {
            "WS1": checklist.cell(row=row, column=checklist_headers["WS1"]).value,
            "WS2": checklist.cell(row=row, column=checklist_headers["WS2"]).value,
        }
        for row in range(2, checklist.max_row + 1)
    }
    assert checklist_rows["NB-ACCESS"]["WS1"] == 1
    assert checklist_rows["NB-ACCESS"]["WS2"] is None
    assert checklist_rows["NB-CACHE"] == {"WS1": 3, "WS2": 0}

    markdown = build_markdown("Fabric Project", aggregate(results), results)
    checklist_header = next(
        line
        for line in markdown.splitlines()
        if line.startswith("| Check ID | Ref | Area |")
    )
    assert checklist_header.endswith("| WS1 | WS2 |")
    checklist_cache_row = next(
        line for line in markdown.splitlines() if line.startswith("| NB-CACHE |")
    )
    assert checklist_cache_row.endswith("| 3.00 | 0.00 |")
    assert "3.00%" not in checklist_cache_row
