from openpyxl import load_workbook

from auditfast.core.enums import Layer, Pillar, Scope, Severity, Status
from auditfast.core.models import CheckResult
from auditfast.core.scoring import aggregate
from auditfast.reporting.excel import build_excel


def _result(**overrides) -> CheckResult:
    base = dict(
        check_id="NB-TEST",
        ref="T.1",
        title="Test finding",
        pillar=Pillar.RELIABILITY,
        status=Status.FAIL,
        score=0,
        evidence="No retry policy is configured.",
        recommendation="Configure a bounded retry.",
        severity=Severity.HIGH,
        workspace="Data Prep",
        layer=Layer.PREP,
        obj="NB_Load",
        scope=Scope.NOTEBOOK,
    )
    base.update(overrides)
    return CheckResult(**base)


def _workbook(tmp_path, results):
    output = tmp_path / "audit.xlsx"
    build_excel(str(output), "Test project", aggregate(results), results)
    return load_workbook(output)


def test_excel_wraps_evidence_and_recommendations(tmp_path):
    recommendation = (
        'Target: notebook "NB_Load" in workspace "Data Prep". '
        "Observed gap: no retry policy. Action: configure a bounded retry. "
        "Verification: re-run the audit."
    )
    result = CheckResult(
        check_id="NB-TEST",
        ref="T.1",
        title="Test finding",
        pillar=Pillar.RELIABILITY,
        status=Status.FAIL,
        score=0,
        evidence="No retry policy is configured.",
        recommendation=recommendation,
        severity=Severity.HIGH,
        workspace="Data Prep",
        layer=Layer.PREP,
        obj="NB_Load",
        scope=Scope.NOTEBOOK,
    )
    output = tmp_path / "audit.xlsx"

    build_excel(str(output), "Test project", aggregate([result]), [result])

    workbook = load_workbook(output)
    findings = workbook["Findings"]
    assert findings["I2"].alignment.wrap_text is True
    assert findings["J2"].alignment.wrap_text is True
    assert findings["J2"].value == recommendation


def test_no_sheet_carries_a_validation_column(tmp_path):
    """The Phase-1 validation flag is not part of the delivered report."""
    workbook = _workbook(tmp_path, [_result()])
    for sheet in workbook.worksheets:
        headers = {
            str(cell.value).strip().casefold()
            for row in sheet.iter_rows()
            for cell in row
            if cell.value is not None
        }
        assert "validation" not in headers, sheet.title
        assert "validated" not in headers, sheet.title
        assert "pending validation" not in headers, sheet.title


def test_summary_and_area_detail_blocks_are_real_tables(tmp_path):
    """Both sheets are stacks of blocks, so each block gets its own table.

    They were the only two sheets shipping as plain cells - every other sheet
    already had a styled table with filter dropdowns and banded rows.
    """
    workbook = _workbook(tmp_path, [_result()])
    summary_tables = workbook["Summary"].tables
    area_tables = workbook["Area Detail"].tables
    assert "SummaryAreaScorecard" in summary_tables
    assert "SummaryWorkspaceScores" in summary_tables
    assert "SummaryCoverage" in summary_tables
    assert "SummaryKeyRisks" in summary_tables
    assert area_tables, "Area Detail should carry at least one block table"


def test_every_populated_sheet_carries_a_table(tmp_path):
    workbook = _workbook(tmp_path, [_result()])
    for sheet in workbook.worksheets:
        if sheet.max_row < 2:
            continue
        assert sheet.tables or sheet.auto_filter.ref, sheet.title


def test_summary_reports_a_percentage_per_workspace(tmp_path):
    """'Which workspace is dragging the score down' is answerable from Summary."""
    results = [
        _result(workspace="Data Prep", score=0, status=Status.FAIL),
        _result(
            check_id="NB-OK", ref="T.2", workspace="Reporting", score=3,
            status=Status.PASS, layer=Layer.REPORTING,
        ),
    ]
    summary = _workbook(tmp_path, results)["Summary"]
    rows = {
        row[0].value: row
        for row in summary.iter_rows()
        if row and isinstance(row[0].value, str)
    }
    assert "Workspace Scores" in rows
    assert "Data Prep" in rows and "Reporting" in rows
    # Column B holds the score as an Excel fraction of 1.
    assert rows["Data Prep"][1].value == 0.0
    assert rows["Reporting"][1].value == 1.0
    # Worst first, so the workspace needing attention is read first.
    assert rows["Data Prep"][0].row < rows["Reporting"][0].row
