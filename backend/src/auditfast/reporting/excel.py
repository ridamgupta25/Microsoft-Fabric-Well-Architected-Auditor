"""SQL-aligned stakeholder Excel report for Fabric Well-Architected audits."""
from __future__ import annotations

from collections import defaultdict
from datetime import date

from ..core.enums import Pillar, Severity, Status
from ..core.scoring import percentage, rating
from ..core.validation import PENDING_LABEL, VALIDATED_LABEL
from .structure import (
    RISK_PROFILE,
    assessment_weight,
    consolidate,
    executive_narrative,
    findings,
    pillar_controls,
    severity_counts,
    strengths,
    validation_counts,
    workspace_control_score,
    workspace_ids,
)


def _pct(pct):
    return None if pct is None else pct / 100


def build_excel(
    path: str,
    project_name: str,
    agg: dict,
    results: list,
    errors: list | None = None,
) -> None:
    """Write the SQL Auditor report flow using consolidated Fabric findings."""
    from openpyxl import Workbook
    from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter, range_boundaries
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = Workbook()
    controls = consolidate(results)
    workspace_id_by_name = workspace_ids(results)
    consolidated_findings = findings(controls)
    consolidated_strengths = strengths(controls)
    validated, total_controls = validation_counts(controls)
    severity_total = severity_counts(controls)
    reported_severities = (
        Severity.CRITICAL,
        Severity.HIGH,
        Severity.MEDIUM,
        Severity.LOW,
    )
    if severity_total[Severity.INFO]:
        reported_severities += (Severity.INFO,)
    pillar_number = {pillar: index for index, pillar in enumerate(Pillar.scored(), start=1)}

    dark_blue = "1F4E78"
    medium_blue = "305496"
    light_blue = "D9EAF7"
    light_green = "E2F0D9"
    light_orange = "FCE4D6"
    light_gray = "F2F2F2"
    banded_fill = PatternFill("solid", fgColor="F7F9FC")
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    partial_fill = PatternFill("solid", fgColor="FFEB9C")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    white = "FFFFFF"
    thin_gray = Side(style="thin", color="D9E1F2")
    header_font = Font(bold=True, color=white)
    title_font = Font(bold=True, size=16, color=white)
    section_font = Font(bold=True, size=12, color=white)

    def style_header(ws, row: int, ncols: int, fill: str = medium_blue) -> None:
        for column in range(1, ncols + 1):
            cell = ws.cell(row=row, column=column)
            cell.font = header_font
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin_gray)

    def style_section(ws, row: int, title: str, ncols: int) -> None:
        ws.cell(row=row, column=1, value=title)
        for column in range(1, ncols + 1):
            cell = ws.cell(row=row, column=column)
            cell.fill = PatternFill("solid", fgColor=dark_blue)
            cell.font = section_font

    def wrap_sheet(ws) -> None:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                    horizontal=cell.alignment.horizontal,
                )

    def auto_fit_columns(ws, max_width: int = 70) -> None:
        for column_index, column_cells in enumerate(ws.columns, start=1):
            column_letter = get_column_letter(column_index)
            content_width = max(
                (
                    max(
                        len(line)
                        for line in (str(cell.value).splitlines() or [""])
                    )
                    for cell in column_cells
                    if cell.value is not None
                ),
                default=0,
            )
            ws.column_dimensions[column_letter].width = min(
                max(content_width + 2, 10),
                max_width,
            )

    def add_styled_table(ws, name: str, ref: str) -> None:
        _, min_row, _, max_row = range_boundaries(ref)
        if min_row == max_row:
            ws.auto_filter.ref = ref
            return
        table = Table(displayName=name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    def add_score_scale(
        ws,
        ref: str,
        *,
        midpoint: float = 0.6,
        maximum: float = 1,
    ) -> None:
        ws.conditional_formatting.add(
            ref,
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="F8696B",
                mid_type="num",
                mid_value=midpoint,
                mid_color="FFEB84",
                end_type="num",
                end_value=maximum,
                end_color="63BE7B",
            ),
        )

    def add_result_formatting(ws) -> None:
        ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
        for value, fill in (
            (Status.PASS.value, pass_fill),
            (Status.PARTIAL.value, partial_fill),
            (Status.FAIL.value, fail_fill),
            (Severity.CRITICAL.value, fail_fill),
            (Severity.HIGH.value, fail_fill),
            (Severity.MEDIUM.value, partial_fill),
            (Severity.LOW.value, pass_fill),
        ):
            ws.conditional_formatting.add(
                ref,
                CellIsRule(
                    operator="equal",
                    formula=[f'"{value}"'],
                    fill=fill,
                ),
            )

    def band_unfilled_rows(ws) -> None:
        for row_number in range(2, ws.max_row + 1, 2):
            row = list(ws.iter_rows(min_row=row_number, max_row=row_number))[0]
            if not any(cell.value is not None for cell in row):
                continue
            for cell in row:
                if cell.fill.fill_type is None:
                    cell.fill = banded_fill

    def append_finding_row(ws, control) -> None:
        likelihood, impact, risk_score, _ = control.risk_profile
        ws.append(
            [
                control.ref,
                control.severity.value,
                risk_score,
                likelihood,
                impact,
                control.impacted_assets,
                control.non_impacted_assets,
                control.not_assessed,
                control.finding,
                control.recommendation,
                "Open",
            ]
        )

    # -- Summary ---------------------------------------------------------------
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"Fabric Audit - {project_name}"
    ws.merge_cells("A1:E1")
    ws["A1"].font = title_font
    ws["A1"].fill = PatternFill("solid", fgColor=dark_blue)
    ws["A1"].alignment = Alignment(horizontal="left")
    ws.freeze_panes = "A3"

    overall_label, overall_emoji = rating(agg["overall"])
    ws["A3"] = "Overall Score"
    ws["B3"] = _pct(agg["overall"])
    ws["B3"].number_format = "0.0%"
    ws["A4"] = "Risk Rating"
    ws["B4"] = f"{overall_emoji} {overall_label}"
    ws["A5"] = "Deployment Mode"
    ws["B5"] = "Microsoft Fabric SaaS"
    for cell in ("A3", "A4", "A5"):
        ws[cell].font = Font(bold=True)

    style_section(ws, 7, "Executive Overview", 5)
    ws["A8"] = executive_narrative(agg, controls)
    ws.merge_cells("A8:E10")
    ws["A8"].alignment = Alignment(vertical="top", wrap_text=True)
    ws["A8"].fill = PatternFill("solid", fgColor=light_blue)

    style_section(ws, 12, "Area Scorecard", 5)
    score_headers = ["#", "Area", "Assessment Weight", "Score", "Rating"]
    for column, value in enumerate(score_headers, start=1):
        ws.cell(row=13, column=column, value=value)
    style_header(ws, 13, len(score_headers))
    row = 14
    for pillar in Pillar.scored():
        pillar_agg = agg["by_pillar"][pillar]
        label, _ = rating(pillar_agg["pct"])
        ws.cell(row=row, column=1, value=pillar_number[pillar])
        ws.cell(row=row, column=2, value=pillar.value)
        ws.cell(row=row, column=3, value=assessment_weight(results, pillar))
        ws.cell(row=row, column=3).number_format = "0.0%"
        ws.cell(row=row, column=4, value=_pct(pillar_agg["pct"]))
        ws.cell(row=row, column=4).number_format = "0.0%"
        ws.cell(row=row, column=5, value=label)
        row += 1
    scorecard_end_row = row - 1

    row += 1
    style_section(ws, row, "Coverage and Validation", 5)
    row += 1
    coverage_headers = ["Measure", "Count", "Measure", "Count", "Rate"]
    for column, value in enumerate(coverage_headers, start=1):
        ws.cell(row=row, column=column, value=value)
    style_header(ws, row, len(coverage_headers))
    control_counts = {status: sum(c.status is status for c in controls) for status in Status}
    coverage_rows = [
        ("Consolidated controls", total_controls, VALIDATED_LABEL, validated),
        ("Passing controls", control_counts[Status.PASS], PENDING_LABEL, total_controls - validated),
        (
            "Failing controls",
            control_counts[Status.FAIL],
            "Validation coverage",
            validated,
        ),
        ("Partial controls", control_counts[Status.PARTIAL], "Asset-level results", len(results)),
        ("Not assessed", control_counts[Status.NA], "Crawl warnings", len(errors or [])),
    ]
    for measure, count, validation, validation_count in coverage_rows:
        row += 1
        ws.cell(row=row, column=1, value=measure)
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=validation)
        ws.cell(row=row, column=4, value=validation_count)
        if validation == "Validation coverage":
            ws.cell(
                row=row,
                column=5,
                value=validated / total_controls if total_controls else 0,
            )
            ws.cell(row=row, column=5).number_format = "0.0%"

    row += 2
    style_section(ws, row, "Key Strengths", 5)
    row += 1
    strength_headers = ["Ref", "Area", "Strength", "Evidence", "Assets Assessed"]
    for column, value in enumerate(strength_headers, start=1):
        ws.cell(row=row, column=column, value=value)
    style_header(ws, row, len(strength_headers), fill="548235")
    for control in consolidated_strengths[:10]:
        row += 1
        ws.cell(row=row, column=1, value=control.ref)
        ws.cell(row=row, column=2, value=control.pillar.value)
        ws.cell(row=row, column=3, value=control.title)
        ws.cell(row=row, column=4, value=control.impacted_evidence)
        ws.cell(row=row, column=5, value=control.assets_assessed)

    row += 2
    style_section(ws, row, "Key Risks", 5)
    row += 1
    risk_headers = ["Ref", "Severity", "Area", "Finding", "Recommendation"]
    for column, value in enumerate(risk_headers, start=1):
        ws.cell(row=row, column=column, value=value)
    style_header(ws, row, len(risk_headers), fill="C65911")
    for control in consolidated_findings[:10]:
        row += 1
        ws.cell(row=row, column=1, value=control.ref)
        ws.cell(row=row, column=2, value=control.severity.value)
        ws.cell(row=row, column=3, value=control.pillar.value)
        ws.cell(row=row, column=4, value=control.finding)
        ws.cell(row=row, column=5, value=control.recommendation)

    # -- Checks sheet ----------------------------------------------------------
    cs = wb.create_sheet("Checks")
    headers = ["Workspace", "Layer role", "Object", "Check ID", "Ref", "Title",
               "Validation", "Pillar", "Status", "Score", "Severity", "Source",
               "Evidence", "Recommendation"]
    cs.append(headers)
    style_header(cs, row=1, ncols=len(headers))
    val_col = headers.index("Validation") + 1
    source_col = headers.index("Source") + 1
    validated_fill = PatternFill("solid", fgColor="C6EFCE")
    pending_fill = PatternFill("solid", fgColor="F2F2F2")
    validated_font = Font(color="006100")
    pending_font = Font(color="808080")
    external_fill = PatternFill("solid", fgColor="FFF2CC")  # Light yellow for external
    external_font = Font(color="9C6500")
    for res in results:
        cs.append([
            res.workspace, res.workspace_role, res.obj, res.check_id, res.ref, res.title,
            validation_label(res.ref),
            res.pillar, res.status.value,
            "" if res.score is None else res.score,
            res.severity.value, res.source, res.evidence, res.recommendation,
        ])
        val_cell = cs.cell(row=cs.max_row, column=val_col)
        if is_validated(res.ref):
            val_cell.fill = validated_fill
            val_cell.font = validated_font
        else:
            val_cell.fill = pending_fill
            val_cell.font = pending_font
        
        # Highlight external checks
        if res.source == "external":
            src_cell = cs.cell(row=cs.max_row, column=source_col)
            src_cell.fill = external_fill
            src_cell.font = external_font
    widths = [22, 16, 22, 14, 8, 34, 16, 22, 10, 7, 12, 12, 46, 52]
    for i, w in enumerate(widths, start=1):
        cs.column_dimensions[cs.cell(row=1, column=i).column_letter].width = w
    cs.freeze_panes = "A2"

    row += 2
    style_section(ws, row, "Remediation Roadmap", 5)
    row += 1
    roadmap_headers = ["Priority", "Findings", "Remediation SLA", "Treatment", "Status"]
    for column, value in enumerate(roadmap_headers, start=1):
        ws.cell(row=row, column=column, value=value)
    style_header(ws, row, len(roadmap_headers))
    for severity in reported_severities:
        row += 1
        *_, sla = RISK_PROFILE[severity]
        ws.cell(row=row, column=1, value=severity.value)
        ws.cell(row=row, column=2, value=severity_total[severity])
        ws.cell(row=row, column=3, value=sla)
        ws.cell(row=row, column=4, value="Mitigate")
        ws.cell(row=row, column=5, value="Open")

    for column, width in {"A": 24, "B": 36, "C": 24, "D": 70, "E": 70}.items():
        ws.column_dimensions[column].width = width
    wrap_sheet(ws)

    # -- Area Detail -----------------------------------------------------------
    area = wb.create_sheet("Area Detail")
    area.freeze_panes = "B2"
    row = 1
    for pillar in Pillar.scored():
        rows = pillar_controls(controls, pillar)
        pillar_results = [result for control in rows for result in control.results]
        score = percentage(pillar_results)
        label, _ = rating(score)
        style_section(
            area,
            row,
            f"Area {pillar_number[pillar]}: {pillar.value}",
            8,
        )
        row += 1
        area.cell(row=row, column=1, value="Area score")
        area.cell(row=row, column=2, value=_pct(score))
        area.cell(row=row, column=2).number_format = "0.0%"
        area.cell(row=row, column=3, value=label)
        area.cell(row=row, column=4, value="Assessment weight")
        area.cell(row=row, column=5, value=assessment_weight(results, pillar))
        area.cell(row=row, column=5).number_format = "0.0%"
        row += 2

        category_map = defaultdict(list)
        for control in rows:
            category_map[control.category].append(control)
        category_header = ["Category", "Score", "Rating", "Validated", "Controls"]
        for column, value in enumerate(category_header, start=1):
            area.cell(row=row, column=column, value=value)
        style_header(area, row, len(category_header))
        for category, category_controls in category_map.items():
            category_results = [
                result for control in category_controls for result in control.results
            ]
            category_score = percentage(category_results)
            category_label, _ = rating(category_score)
            row += 1
            area.cell(row=row, column=1, value=category)
            area.cell(row=row, column=2, value=_pct(category_score))
            area.cell(row=row, column=2).number_format = "0.0%"
            area.cell(row=row, column=3, value=category_label)
            area.cell(
                row=row,
                column=4,
                value=sum(control.validation == VALIDATED_LABEL for control in category_controls),
            )
            area.cell(row=row, column=5, value=len(category_controls))
        row += 2

        pillar_strengths = strengths(rows)
        area.cell(row=row, column=1, value=f"Strengths ({len(pillar_strengths)})")
        style_header(area, row, 8, fill="548235")
        row += 1
        strength_detail_headers = [
            "Ref",
            "Strength",
            "Severity",
            "Score",
            "Assets",
            "Validation",
            "Evidence",
            "Status",
        ]
        for column, value in enumerate(strength_detail_headers, start=1):
            area.cell(row=row, column=column, value=value)
        style_header(area, row, len(strength_detail_headers), fill="70AD47"        )
        for control in pillar_strengths:
            row += 1
            values = [
                control.ref,
                control.title,
                control.severity.value,
                control.score_summary,
                control.assets_assessed,
                control.validation,
                control.impacted_evidence,
                control.status.value,
            ]
            for column, value in enumerate(values, start=1):
                area.cell(row=row, column=column, value=value)
        row += 2

        pillar_findings = findings(rows)
        area.cell(row=row, column=1, value=f"Findings ({len(pillar_findings)})")
        style_header(area, row, 8, fill="C65911")
        row += 1
        finding_detail_headers = [
            "Ref",
            "Finding",
            "Severity",
            "Score",
            "Impacted Assets",
            "Non-Impacted Assets",
            "Not Assessed / Reason",
            "Recommendation",
        ]
        for column, value in enumerate(finding_detail_headers, start=1):
            area.cell(row=row, column=column, value=value)
        style_header(area, row, len(finding_detail_headers))
        for control in pillar_findings:
            row += 1
            values = [
                control.ref,
                control.finding,
                control.severity.value,
                control.score_summary,
                control.impacted_assets,
                control.non_impacted_assets,
                control.not_assessed,
                control.recommendation,
            ]
            for column, value in enumerate(values, start=1):
                area.cell(row=row, column=column, value=value)
        row += 3

    area_widths = [18, 70, 18, 18, 36, 36, 70, 80]
    for index, width in enumerate(area_widths, start=1):
        area.column_dimensions[area.cell(row=1, column=index).column_letter].width = width
    wrap_sheet(area)

    # -- Checklist -------------------------------------------------------------
    checklist = wb.create_sheet("Checklist")
    checklist_headers = [
        "Check ID",
        "Ref",
        "Area",
        "Category",
        "Check Description",
        "ProducedBy",
        "Confidence",
        "Severity",
        "Validation",
        *workspace_id_by_name.values(),
    ]
    checklist.append(checklist_headers)
    style_header(checklist, 1, len(checklist_headers))
    workspace_columns = {
        workspace: checklist_headers.index(workspace_id) + 1
        for workspace, workspace_id in workspace_id_by_name.items()
    }
    for control in controls:
        checklist.append(
            [
                control.check_id,
                control.ref,
                pillar_number.get(control.pillar, ""),
                control.category,
                control.title,
                "Rule-based",
                "Deterministic",
                control.severity.value,
                control.validation,
                *(
                    workspace_control_score(control, workspace)
                    for workspace in workspace_id_by_name
                ),
            ]
        )
        for workspace, column in workspace_columns.items():
            value = workspace_control_score(control, workspace)
            if isinstance(value, float):
                checklist.cell(
                    row=checklist.max_row,
                    column=column,
                    value=value,
                )
                checklist.cell(row=checklist.max_row, column=column).number_format = "0.00"
    checklist.freeze_panes = "J2"
    last_checklist_column = checklist.cell(
        row=1,
        column=len(checklist_headers),
    ).column_letter
    add_styled_table(
        checklist,
        "ChecklistTable",
        f"A1:{last_checklist_column}{checklist.max_row}",
    )
    checklist_widths = [
        18,
        10,
        8,
        28,
        58,
        14,
        16,
        12,
        18,
        *(14 for _ in workspace_id_by_name),
    ]
    for index, width in enumerate(checklist_widths, start=1):
        checklist.column_dimensions[
            checklist.cell(row=1, column=index).column_letter
        ].width = width
    wrap_sheet(checklist)

    # -- Findings --------------------------------------------------------------
    finding_sheet = wb.create_sheet("Findings")
    finding_headers = [
        "Ref",
        "Severity",
        "Risk",
        "Likelihood",
        "Impact",
        "Impacted Assets",
        "Non-Impacted Assets",
        "Not Assessed / Reason",
        "Finding",
        "Recommendation",
        "Status",
    ]
    finding_sheet.append(finding_headers)
    style_header(finding_sheet, 1, len(finding_headers))
    for control in consolidated_findings:
        append_finding_row(finding_sheet, control)
    finding_sheet.freeze_panes = "F2"
    add_styled_table(
        finding_sheet,
        "FindingsTable",
        f"A1:K{finding_sheet.max_row}",
    )
    finding_widths = [10, 12, 10, 12, 10, 36, 36, 60, 80, 90, 12]
    for index, width in enumerate(finding_widths, start=1):
        finding_sheet.column_dimensions[
            finding_sheet.cell(row=1, column=index).column_letter
        ].width = width
    wrap_sheet(finding_sheet)

    # -- Risk Register ---------------------------------------------------------
    risk = wb.create_sheet("Risk Register")
    risk["A1"] = f"Risk Register - Fabric Audit - {project_name}"
    risk.merge_cells("A1:Z1")
    risk["A1"].font = title_font
    risk["A1"].fill = PatternFill("solid", fgColor=dark_blue)
    risk["A2"] = "Generated"
    risk["B2"] = date.today().isoformat()
    risk["D2"] = "Overall score"
    risk["E2"] = _pct(agg["overall"])
    risk["E2"].number_format = "0.0%"
    risk["G2"] = "Overall rating"
    risk["H2"] = overall_label
    risk["A4"] = "Risk Summary"
    risk["G4"] = (
        "Assign an owner and treatment, confirm the target date, update status and "
        "closure evidence, and mark verification only after an independent re-test. "
        "Repeated asset failures are consolidated into one control-level risk."
    )
    risk.merge_cells("G4:O8")
    risk["G4"].fill = PatternFill("solid", fgColor=light_blue)
    risk["G4"].alignment = Alignment(vertical="top", wrap_text=True)
    risk_summary_headers = ["Severity", "Count", "% of findings", "Remediation SLA"]
    for column, value in enumerate(risk_summary_headers, start=1):
        risk.cell(row=5, column=column, value=value)
    style_header(risk, 5, len(risk_summary_headers))
    row = 6
    for severity in reported_severities:
        *_, sla = RISK_PROFILE[severity]
        count = severity_total[severity]
        risk.cell(row=row, column=1, value=severity.value)
        risk.cell(row=row, column=2, value=count)
        risk.cell(
            row=row,
            column=3,
            value=count / len(consolidated_findings) if consolidated_findings else 0,
        )
        risk.cell(row=row, column=3).number_format = "0.0%"
        risk.cell(row=row, column=4, value=sla)
        row += 1
    risk.cell(row=row, column=1, value="Total")
    risk.cell(row=row, column=2, value=len(consolidated_findings))
    risk.cell(row=row, column=3, value=1 if consolidated_findings else 0)
    risk.cell(row=row, column=3).number_format = "0.0%"

    risk_header_row = row + 1
    risk_headers = [
        "Risk ID",
        "Audit Phase",
        "Area #",
        "Area",
        "Category",
        "Checklist Ref",
        "Check ID",
        "Finding / Current Evidence",
        "Scope",
        "Impacted Assets",
        "Non-Impacted Assets",
        "Not Assessed / Reason",
        "Severity",
        "Likelihood",
        "Impact",
        "Risk Score",
        "Remediation SLA",
        "Recommendation",
        "Owner",
        "Target Date",
        "Treatment",
        "Status",
        "Actual Closure Date",
        "Closure Evidence",
        "Verification Status",
        "Notes",
    ]
    for column, value in enumerate(risk_headers, start=1):
        risk.cell(row=risk_header_row, column=column, value=value)
    style_header(risk, risk_header_row, len(risk_headers))
    for index, control in enumerate(consolidated_findings, start=1):
        likelihood, impact, risk_score, sla = control.risk_profile
        risk.append(
            [
                f"R-{index:03d}",
                "Automated assessment",
                pillar_number.get(control.pillar, ""),
                control.pillar.value,
                control.category,
                control.ref,
                control.check_id,
                control.impacted_evidence,
                control.scopes,
                control.impacted_assets,
                control.non_impacted_assets,
                control.not_assessed,
                control.severity.value,
                likelihood,
                impact,
                risk_score,
                sla,
                control.recommendation,
                "",
                "",
                "Mitigate",
                "Open",
                "",
                "",
                "Not verified",
                "",
            ]
        )
    risk.freeze_panes = f"I{risk_header_row + 1}"
    add_styled_table(
        risk,
        "RiskRegisterTable",
        f"A{risk_header_row}:Z{risk.max_row}",
    )
    risk_widths = [
        12,
        22,
        10,
        28,
        30,
        14,
        18,
        80,
        20,
        40,
        40,
        70,
        12,
        12,
        10,
        12,
        20,
        90,
        24,
        16,
        16,
        14,
        20,
        50,
        20,
        50,
    ]
    for index, width in enumerate(risk_widths, start=1):
        risk.column_dimensions[
            risk.cell(row=risk_header_row, column=index).column_letter
        ].width = width
    wrap_sheet(risk)

    # -- Inventory -------------------------------------------------------------
    inventory = wb.create_sheet("Invent")
    inventory_headers = [
        "Workspace ID",
        "Name",
        "Layer Role",
        "Objects Assessed",
        "Asset-Level Results",
        "Consolidated Controls",
    ]
    inventory.append(inventory_headers)
    style_header(inventory, 1, len(inventory_headers))
    for workspace in sorted({result.workspace for result in results if result.workspace}):
        workspace_results = [result for result in results if result.workspace == workspace]
        inventory.append(
            [
                workspace_id_by_name[workspace],
                workspace,
                workspace_results[0].workspace_role,
                len({result.obj for result in workspace_results if result.obj}),
                len(workspace_results),
                len({(result.check_id, result.ref) for result in workspace_results}),
            ]
        )
    inventory.freeze_panes = "C2"
    add_styled_table(
        inventory,
        "InventoryTable",
        f"A1:F{inventory.max_row}",
    )
    inventory_widths = [14, 42, 22, 20, 22, 24]
    for index, width in enumerate(inventory_widths, start=1):
        inventory.column_dimensions[
            inventory.cell(row=1, column=index).column_letter
        ].width = width
    wrap_sheet(inventory)

    # Light semantic fills make the stakeholder hierarchy visible without
    # changing any deterministic values.
    for sheet in (ws, area, checklist, finding_sheet, risk, inventory):
        sheet.sheet_view.showGridLines = False
    ws["A3"].fill = PatternFill("solid", fgColor=light_gray)
    ws["A4"].fill = PatternFill("solid", fgColor=light_orange)
    ws["A5"].fill = PatternFill("solid", fgColor=light_green)

    add_score_scale(ws, "B3")
    add_score_scale(ws, f"D14:D{scorecard_end_row}")
    add_score_scale(area, f"B1:B{area.max_row}")
    add_score_scale(risk, "E2")
    if workspace_id_by_name:
        first_workspace_column = len(checklist_headers) - len(workspace_id_by_name) + 1
        first_workspace_letter = checklist.cell(
            row=1,
            column=first_workspace_column,
        ).column_letter
        add_score_scale(
            checklist,
            f"{first_workspace_letter}2:{last_checklist_column}{checklist.max_row}",
            midpoint=1.5,
            maximum=3,
        )

    for sheet in (ws, area, checklist, finding_sheet, risk, inventory):
        add_result_formatting(sheet)
    band_unfilled_rows(ws)
    band_unfilled_rows(area)

    for sheet, max_width in (
        (ws, 70),
        (area, 70),
        (checklist, 60),
        (finding_sheet, 70),
        (risk, 70),
        (inventory, 45),
    ):
        auto_fit_columns(sheet, max_width=max_width)

    wb.save(path)
